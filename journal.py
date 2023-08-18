import tkinter
from tkinter import ttk, messagebox, filedialog
from tkcalendar import DateEntry
from datetime import date, datetime, timedelta
import journal_voucher
import sqlcommand
import invoice
import openpyxl
import csv


class Journal(tkinter.Toplevel):

    def __init__(self, parent, date_from, date_to, accounts_list):
        super().__init__(parent)

        self.accounts_list = accounts_list
        self.date_from = date_from
        self.date_to = date_to

        self.title(f"Journal from {self.date_from} to {self.date_to}")
        self.state('zoomed')

        self.journal_frame = tkinter.Frame(self)
        self.journal_frame.pack(fill='both', expand=True)

        self.journal_tree_scroll_y = tkinter.Scrollbar(self.journal_frame, orient="vertical")
        self.journal_tree_scroll_y.pack(side='right', fill='y')

        self.journal_treeview()

        # Create Options Frame
        self.journal_options = tkinter.Frame(self)
        self.journal_options.pack(side='right')

        self.view_voucher_button = tkinter.Button(self.journal_options, text="View voucher",
                                                  command=self.view_voucher, pady=15, padx=15, width=20)
        self.view_voucher_button.grid(column=0, row=0)

        self.view_document_button = tkinter.Button(self.journal_options, text="View original document",
                                                  command=self.view_document, pady=15, padx=15, width=20)
        self.view_document_button.grid(column=1, row=0)

        self.export_to_excel_button = tkinter.Button(self.journal_options, text="Export to XLS file",
                                                     command=lambda: self.export_to_excel(self.journal_tree),
                                                     pady=15, padx=15, width=20)
        self.export_to_excel_button.grid(column=2, row=0)

        self.export_to_csv_button = tkinter.Button(self.journal_options, text="Export to CSV file",
                                                   command=lambda: self.export_to_csv(self.journal_tree),
                                                   pady=15, padx=15, width=20)
        self.export_to_csv_button.grid(column=3, row=0)

    def journal_treeview(self):
        """Create chart of accounts treeview"""
        # Connect to sql database to add account
        self.sql_command = f'''
            SELECT no_journal, journal_entry_date, doc_type, doc_date, description,
            COALESCE(journal.dt_account_num, '') AS dt_account_num,
            COALESCE(dt.account_name, '') AS dt_account_name,
            COALESCE(journal.ct_account_num, '') AS ct_account_num,
            COALESCE(ct.account_name, '') AS ct_account_name,
            TO_CHAR(COALESCE(CAST(amount AS DECIMAL(1000, 2)),0), 'fm999G999G999G990.00') amount,
            COALESCE(journal.cust_id, 0) AS cust_id, COALESCE(customers.cust_name,'') AS customer_name,
            doc_number, int_doc_number, COALESCE(vat_id,'') AS vat_id, vat_date, COALESCE(status, '') AS status, 
            cor_orig_doc_number
            FROM journal
            LEFT JOIN customers
            ON journal.cust_id = customers.cust_id
            LEFT JOIN chart_of_accounts AS dt
            ON journal.dt_account_num = dt.account_num
            LEFT JOIN chart_of_accounts AS ct
            ON journal.ct_account_num = ct.account_num
            WHERE doc_date BETWEEN '{self.date_from}' AND '{self.date_to}' AND
            (dt_account_num ILIKE ANY(ARRAY{self.accounts_list}) OR ct_account_num ILIKE ANY(ARRAY{self.accounts_list}))
            '''
        self.order_by = ' ORDER BY doc_date ASC'

        conn = sqlcommand.SqlCommand()
        execute = conn.sql_execute(self.sql_command + self.order_by)

        column_names = execute['column_names']

        self.journal_tree = ttk.Treeview(self.journal_frame, columns=column_names, show='headings', height=35,
                                         selectmode='browse', yscrollcommand=self.journal_tree_scroll_y.set)
        self.journal_tree_scroll_y.config(command=self.journal_tree.yview)

        column = 0
        for _ in column_names:
            self.journal_tree.heading(column_names[column], text=f'{column_names[column]}')
            column += 1

        # Adjust column width
        self.journal_tree.column(column_names[0], width=0)
        self.journal_tree.column(column_names[1], width=170)
        self.journal_tree.column(column_names[2], width=100)
        self.journal_tree.column(column_names[3], width=100)
        self.journal_tree.column(column_names[4], width=170)
        self.journal_tree.column(column_names[5], width=100, anchor='e')
        self.journal_tree.column(column_names[6], width=170)
        self.journal_tree.column(column_names[7], width=100, anchor='e')
        self.journal_tree.column(column_names[8], width=170)
        self.journal_tree.column(column_names[9], width=130, anchor='e')
        self.journal_tree.column(column_names[10], width=100, anchor='e')
        self.journal_tree.column(column_names[11], width=170)
        self.journal_tree.column(column_names[12], width=140)
        self.journal_tree.column(column_names[13], width=140)
        self.journal_tree.column(column_names[14], width=100)
        self.journal_tree.column(column_names[15], width=100)
        self.journal_tree.column(column_names[16], width=100)

        self.journal_tree["displaycolumns"] = (column_names[1:-1])

        self.journal_tree.pack(fill='both', expand=True)

        records = execute['records']

        record = 0
        for _ in records:
            self.journal_tree.insert('', tkinter.END, values=records[record])
            record += 1

    def selected_document(self):
        """Take selected document from tree"""
        selected = self.journal_tree.focus()
        self.selected_doc = self.journal_tree.item(selected).get('values')
        return self.selected_doc

    def view_voucher(self):
        """Open window of journal voucher to view"""
        try:
            journal_type = 'Opening balance' if self.selected_document()[2] == 'Opening balance' else 'Journal voucher'
            doc_type = self.selected_document()[2]
            doc_date = self.selected_document()[3]
            int_doc_number = self.selected_document()[13]
            journal_voucher.ViewJournalVoucher(self, journal_type, doc_type, doc_date, int_doc_number)
        except:
            tkinter.messagebox.showerror(title="ERROR", message="Select document", parent=self)

    def view_document(self):
        """Open window of original document to view"""
        doc_type = self.selected_document()[2]
        doc_date = self.selected_document()[3]
        journal_entry_date = self.selected_document()[1]
        doc_number = self.selected_document()[12]
        cust_id = self.selected_document()[10]
        cust_name = self.selected_document()[11]
        orig_doc_number = self.selected_document()[17]
        int_doc_number = self.selected_document()[13]

        if doc_type == 'purchase costs':
            doc_type = 'Cost Purchase Invoice'
            invoice.ViewInvoice(self, doc_type, journal_entry_date, doc_number, cust_id, cust_name)
        elif doc_type == 'sale service':
            doc_type = 'Service Sales Invoice'
            invoice.ViewInvoice(self, doc_type, journal_entry_date, doc_number, cust_id, cust_name)

        elif doc_type == '[COR] purchase costs':
            doc_type = 'Cost Purchase Invoice'
            invoice.ViewCorrectingInvoice(self, doc_type, journal_entry_date, doc_number, cust_id, cust_name,
                                          orig_doc_number, 'view')
        elif doc_type == '[COR] sale service':
            doc_type = 'Service Sales Invoice'
            invoice.ViewCorrectingInvoice(self, doc_type, journal_entry_date, doc_number, cust_id, cust_name,
                                          orig_doc_number, 'view')
        else:
            journal_voucher.ViewJournalVoucher(self, 'Journal voucher', doc_type, doc_date, int_doc_number)

    def export_to_excel(self, treeview):
        """Export treeview to XLS file"""
        # A dialog box for selecting the location and name of the file
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")],
                                                 parent=self)
        if file_path:
            # Creating a new Excel file
            workbook = openpyxl.Workbook()
            worksheet = workbook.active

            # Getting column headers from treeview
            headers = [treeview.heading(column)["text"] for column in treeview["columns"]]
            # Save column headers to the first row of the worksheet
            for col_idx, header in enumerate(headers):
                worksheet.cell(row=1, column=col_idx + 1).value = header

            # Saving data to the next rows of the spreadsheet
            data = []
            for row_id in treeview.get_children():
                item_data = []
                for col_idx, value in enumerate(treeview.item(row_id)["values"]):
                    item_data.append(value)
                data.append(item_data)

            # Saving data to the next rows of the spreadsheet
            for row_idx, row_data in enumerate(data):
                for col_idx, value in enumerate(row_data):
                    worksheet.cell(row=row_idx + 2, column=col_idx + 1).value = value

            # Saving an Excel file
            workbook.save(file_path)

    def export_to_csv(self, treeview):
        """Export treeview to CSV file"""
        # A dialog box for selecting the location and name of the file
        file_path = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("CSV files", "*.csv")],
                                                 parent=self)
        if file_path:
            # Retrieving data from treeview
            data = []
            for row_id in treeview.get_children():
                item_data = []
                for value in treeview.item(row_id)["values"]:
                    item_data.append(value)
                data.append(item_data)

            # Saving data to a CSV file
            with open(file_path, "w", newline="") as csvfile:
                writer = csv.writer(csvfile)
                writer.writerow(
                    [treeview.heading(column)["text"] for column in
                     treeview["columns"]])
                writer.writerows(data)


class SetPeriod(tkinter.Toplevel):

    def __init__(self, parent):
        super().__init__(parent)

        self.title("Journal - Set period")
        self.config(padx=15, pady=15)
        self.resizable(False, False)

        today = date.today()
        first_day_of_month = date(today.year, today.month, 1)

        self.date_from_label = tkinter.Label(self, text="Date from")
        self.date_from_label.grid(column=0, row=0)
        self.date_from_entry = DateEntry(self, selectmode='day', locale='en_US', date_pattern='dd-mm-y')
        self.date_from_entry.set_date(first_day_of_month)
        self.date_from_entry.grid(column=1, row=0)

        today = date.today()
        next_month = today.replace(day=28) + timedelta(days=4)
        end_of_month = next_month - timedelta(days=next_month.day)

        self.date_to_label = tkinter.Label(self, text="Date to")
        self.date_to_label.grid(column=0, row=1)
        self.date_to_entry = DateEntry(self, selectmode='day', locale='en_US', date_pattern='dd-mm-y')
        self.date_to_entry.set_date(end_of_month)
        self.date_to_entry.grid(column=1, row=1)

        self.account_label = tkinter.Label(self, text="Account")
        self.account_label.grid(column=0, row=2)
        self.account_entry = tkinter.Entry(self, width=20)
        self.account_entry.grid(column=1, row=2)

        self.set_button = tkinter.Button(self, text='Ok', command=self.set_button_command, width=10)
        self.set_button.grid(column=1, row=3)

    def set_button_command(self):
        accounts = self.account_entry.get().replace(" ", "")
        accounts_list = accounts.split(',')
        for item in accounts_list:
            accounts_list[accounts_list.index(item)] = item+'%'

        date_from_str = self.date_from_entry.get()
        date_obj = datetime.strptime(date_from_str, "%d-%m-%Y")
        date_from = date_obj.strftime("%Y-%m-%d")

        date_to_str = self.date_to_entry.get()
        date_obj = datetime.strptime(date_to_str, "%d-%m-%Y")
        date_to = date_obj.strftime("%Y-%m-%d")

        year = int(date_obj.strftime("%Y"))

        self.first_day_of_year = date(year, 1, 1)

        date_format = "%d-%m-%Y"
        date_1 = datetime.strptime(date_from_str, date_format).date()
        date_2 = datetime.strptime(date_to_str, date_format).date()

        if date_1 > date_2:
            tkinter.messagebox.showerror(title="ERROR", message="Date from cannot be earlier than date to", parent=self)
        elif date_1.year != date_2.year:
            tkinter.messagebox.showerror(title="ERROR", message="Dates must be from one fiscal year", parent=self)
        else:
            Journal(self, date_from, date_to, accounts_list)
            self.withdraw()
