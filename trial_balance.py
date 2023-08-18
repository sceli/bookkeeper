import tkinter
from tkinter import ttk, messagebox, filedialog
from tkcalendar import DateEntry
from datetime import date, datetime, timedelta
import sqlcommand
import openpyxl
import csv


class TrialBalance(tkinter.Toplevel):

    def __init__(self, parent, first_day_of_year, date_from, date_to):
        super().__init__(parent)

        self.first_day_of_year = first_day_of_year
        self.date_from = date_from
        self.date_to = date_to

        self.title(f"Trial Balance from {self.date_from} to {self.date_to}")
        self.resizable(False, False)

        self.trial_balance_frame = tkinter.Frame(self)
        self.trial_balance_frame.grid(column=0, row=0)

        self.trial_balance_tree_scroll = tkinter.Scrollbar(self.trial_balance_frame)
        self.trial_balance_tree_scroll.grid(column=1, row=0, sticky='ns')

        self.trial_balance_treeview()

        # Create Options Frame
        self.trial_balance_options = tkinter.Frame(self)
        self.trial_balance_options.grid(column=0, row=1, sticky='e')

        self.export_to_excel_button = tkinter.Button(self.trial_balance_options, text="Export to XLS file",
                                                     command=lambda: self.export_to_excel(self.trial_balance_tree),
                                                     pady=15, padx=15, width=20)
        self.export_to_excel_button.grid(column=0, row=0)

        self.export_to_csv_button = tkinter.Button(self.trial_balance_options, text="Export to CSV file",
                                                   command=lambda: self.export_to_csv(self.trial_balance_tree),
                                                   pady=15, padx=15, width=20)
        self.export_to_csv_button.grid(column=1, row=0)

    def trial_balance_treeview(self):
        """Create chart of accounts treeview"""
        # Connect to sql database to add account
        self.sql_command = f'''
            SELECT
                subquery.account_num,
                chart_of_accounts.account_name,
                COALESCE(TO_CHAR(SUM(CASE WHEN subquery.doc_type = 'Opening balance' 
                AND subquery.account_type = 'dt_account_num' 
                THEN subquery.amount ELSE 0 END), 'FM999G999G999G990.00'), '0.00') AS dt_open_bal,
                COALESCE(TO_CHAR(SUM(CASE WHEN subquery.doc_type = 'Opening balance' 
                AND subquery.account_type = 'ct_account_num' 
                THEN subquery.amount ELSE 0 END), 'FM999G999G999G990.00'), '0.00') AS ct_open_bal,
                COALESCE(TO_CHAR(SUM(CASE WHEN subquery.doc_type != 'Opening balance' 
                AND subquery.account_type = 'dt_period' 
                THEN subquery.amount ELSE 0 END), 'FM999G999G999G990.00'), '0.00') AS dt_period,
                COALESCE(TO_CHAR(SUM(CASE WHEN subquery.doc_type != 'Opening balance' 
                AND subquery.account_type = 'ct_period' 
                THEN subquery.amount ELSE 0 END), 'FM999G999G999G990.00'), '0.00') AS ct_period,
                COALESCE(TO_CHAR(dt_amounts_full.dt_amount, 'FM999G999G999G990.00'), '0.00') AS dt_cumulatively,
                COALESCE(TO_CHAR(ct_amounts_full.ct_amount, 'FM999G999G999G990.00'), '0.00') AS ct_cumulatively,
                COALESCE(TO_CHAR(COALESCE(dt_amounts_full.dt_amount, 0) - 
                COALESCE(ct_amounts_full.ct_amount, 0), 'FM999G999G999G990.00'), '0.00') AS balance
            FROM
                (
                    SELECT dt_account_num AS account_num, amount, 'dt_period' AS account_type, doc_type, doc_date 
                    FROM journal WHERE doc_date BETWEEN '{self.date_from}' AND '{self.date_to}'
                    UNION ALL
                    SELECT ct_account_num AS account_num, amount, 'ct_period' AS account_type, doc_type, doc_date 
                    FROM journal WHERE doc_date BETWEEN '{self.date_from}' AND '{self.date_to}'
                    UNION ALL
                    SELECT dt_account_num AS account_num, amount, 'dt_account_num' AS account_type, doc_type, doc_date 
                    FROM journal WHERE doc_date BETWEEN '{self.first_day_of_year}' AND '{self.date_to}'
                    UNION ALL
                    SELECT ct_account_num AS account_num, amount, 'ct_account_num' AS account_type, doc_type, doc_date 
                    FROM journal WHERE doc_date BETWEEN '{self.first_day_of_year}' AND '{self.date_to}'
                ) AS subquery
            JOIN
                chart_of_accounts ON subquery.account_num = chart_of_accounts.account_num
            LEFT JOIN
                (
                    SELECT dt_account_num, SUM(amount) AS dt_amount FROM journal WHERE doc_date 
                    BETWEEN '{self.first_day_of_year}' AND '{self.date_to}' GROUP BY dt_account_num
                ) AS dt_amounts_full ON subquery.account_num = dt_amounts_full.dt_account_num
            LEFT JOIN
                (
                    SELECT ct_account_num, SUM(amount) AS ct_amount FROM journal WHERE doc_date 
                    BETWEEN '{self.first_day_of_year}' AND '{self.date_to}' GROUP BY ct_account_num
                ) AS ct_amounts_full ON subquery.account_num = ct_amounts_full.ct_account_num
            GROUP BY
                subquery.account_num,
                chart_of_accounts.account_name,
                dt_amounts_full.dt_amount,
                ct_amounts_full.ct_amount
            ORDER BY
                subquery.account_num;
                    '''

        conn = sqlcommand.SqlCommand()
        execute = conn.sql_execute(self.sql_command)

        column_names = execute['column_names']

        self.trial_balance_tree = ttk.Treeview(self.trial_balance_frame, columns=column_names, show='headings',
                                               height=35, selectmode='browse',
                                               yscrollcommand=self.trial_balance_tree_scroll.set)
        self.trial_balance_tree_scroll.config(command=self.trial_balance_tree.yview)

        column = 0
        for _ in column_names:
            self.trial_balance_tree.heading(column_names[column], text=f'{column_names[column]}')
            column += 1

        # Adjust column width
        self.trial_balance_tree.column(column_names[0], width=100)
        self.trial_balance_tree.column(column_names[1], width=200)
        self.trial_balance_tree.column(column_names[2], width=130, anchor='e')
        self.trial_balance_tree.column(column_names[3], width=130, anchor='e')
        self.trial_balance_tree.column(column_names[4], width=130, anchor='e')
        self.trial_balance_tree.column(column_names[5], width=130, anchor='e')
        self.trial_balance_tree.column(column_names[6], width=130, anchor='e')
        self.trial_balance_tree.column(column_names[7], width=130, anchor='e')
        self.trial_balance_tree.column(column_names[8], width=130, anchor='e')

        self.trial_balance_tree.grid(row=0, column=0)

        records = execute['records']

        # Calculate sum for all columns
        dt_open_bal_sum = 0
        ct_open_bal_sum = 0
        dt_period_sum = 0
        ct_period_sum = 0
        dt_cumulatively_sum = 0
        ct_cumulatively_sum = 0
        balance_sum = 0

        for item in records:
            dt_open_bal_sum += float(str(item[2]).replace('\xa0', ''))
            ct_open_bal_sum += float(str(item[3]).replace('\xa0', ''))
            dt_period_sum += float(str(item[4]).replace('\xa0', ''))
            ct_period_sum += float(str(item[5]).replace('\xa0', ''))
            dt_cumulatively_sum += float(str(item[6]).replace('\xa0', ''))
            ct_cumulatively_sum += float(str(item[7]).replace('\xa0', ''))
            balance_sum += float(str(item[8]).replace('\xa0', ''))

        sum_list = ['', 'TOTAL', '{:,.2f}'.format(dt_open_bal_sum).replace(',', ' '),
                    '{:,.2f}'.format(ct_open_bal_sum).replace(',', ' '),
                    '{:,.2f}'.format(dt_period_sum).replace(',', ' '),
                    '{:,.2f}'.format(ct_period_sum).replace(',', ' '),
                    '{:,.2f}'.format(dt_cumulatively_sum).replace(',', ' '),
                    '{:,.2f}'.format(ct_cumulatively_sum).replace(',', ' '),
                    '{:,.2f}'.format(balance_sum).replace(',', ' ')]

        self.trial_balance_tree.tag_configure('sum', background='light grey')

        record = 0
        for _ in records:
            self.trial_balance_tree.insert('', tkinter.END, values=records[record])
            record += 1
        self.trial_balance_tree.insert('', tkinter.END, values=sum_list, tags='sum')

    def set_period_button_command(self):
        SetPeriod(self)

    def export_to_excel(self, treeview):
        """Export treeview to XLS file"""
        # A dialog box for selecting the location and name of the file
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")],
                                                 parent=self)
        if file_path:
            # Creating a new Excel file
            workbook = openpyxl.Workbook()
            worksheet = workbook.active

            # Getting column headers from treeview
            headers = [treeview.heading(column)["text"] for column in treeview["columns"]]
            # Save column headers to the first row of the worksheet
            for col_idx, header in enumerate(headers):
                worksheet.cell(row=1, column=col_idx + 1).value = header

            # Saving data to the next rows of the spreadsheet
            data = []
            for row_id in treeview.get_children():
                item_data = []
                for col_idx, value in enumerate(treeview.item(row_id)["values"]):
                    item_data.append(value)
                data.append(item_data)

            # Saving data to the next rows of the spreadsheet
            for row_idx, row_data in enumerate(data):
                for col_idx, value in enumerate(row_data):
                    worksheet.cell(row=row_idx + 2, column=col_idx + 1).value = value

            # Saving an Excel file
            workbook.save(file_path)

    def export_to_csv(self, treeview):
        """Export treeview to CSV file"""
        # A dialog box for selecting the location and name of the file
        file_path = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("CSV files", "*.csv")],
                                                 parent=self)
        if file_path:
            # Retrieving data from treeview
            data = []
            for row_id in treeview.get_children():
                item_data = []
                for value in treeview.item(row_id)["values"]:
                    item_data.append(value)
                data.append(item_data)

            # Saving data to a CSV file
            with open(file_path, "w", newline="") as csvfile:
                writer = csv.writer(csvfile)
                writer.writerow(
                    [treeview.heading(column)["text"] for column in
                     treeview["columns"]])
                writer.writerows(data)


class SetPeriod(tkinter.Toplevel):

    def __init__(self, parent):
        super().__init__(parent)

        self.title("Trial Balance - Set period")
        self.config(padx=15, pady=15)
        self.resizable(False, False)

        today = date.today()
        first_day_of_month = date(today.year, today.month, 1)

        self.date_from_label = tkinter.Label(self, text="Date from")
        self.date_from_label.grid(column=0, row=0)
        self.date_from_entry = DateEntry(self, selectmode='day', locale='en_US', date_pattern='dd-mm-y')
        self.date_from_entry.set_date(first_day_of_month)
        self.date_from_entry.grid(column=1, row=0)

        today = date.today()
        next_month = today.replace(day=28) + timedelta(days=4)
        end_of_month = next_month - timedelta(days=next_month.day)

        self.date_to_label = tkinter.Label(self, text="Date to")
        self.date_to_label.grid(column=0, row=1)
        self.date_to_entry = DateEntry(self, selectmode='day', locale='en_US', date_pattern='dd-mm-y')
        self.date_to_entry.set_date(end_of_month)
        self.date_to_entry.grid(column=1, row=1)

        self.set_button = tkinter.Button(self, text='Ok', command=self.set_button_command, width=10)
        self.set_button.grid(column=1, row=2)

    def set_button_command(self):
        date_from_str = self.date_from_entry.get()
        date_obj = datetime.strptime(date_from_str, "%d-%m-%Y")
        date_from = date_obj.strftime("%Y-%m-%d")

        date_to_str = self.date_to_entry.get()
        date_obj = datetime.strptime(date_to_str, "%d-%m-%Y")
        date_to = date_obj.strftime("%Y-%m-%d")

        year = int(date_obj.strftime("%Y"))

        self.first_day_of_year = date(year, 1, 1)

        date_format = "%d-%m-%Y"
        date_1 = datetime.strptime(date_from_str, date_format).date()
        date_2 = datetime.strptime(date_to_str, date_format).date()

        if date_1 > date_2:
            tkinter.messagebox.showerror(title="ERROR", message="Date from cannot be earlier than date to", parent=self)
        elif date_1.year != date_2.year:
            tkinter.messagebox.showerror(title="ERROR", message="Dates must be from one fiscal year", parent=self)
        else:
            TrialBalance(self, self.first_day_of_year, date_from, date_to)
            self.withdraw()
